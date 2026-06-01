from __future__ import annotations

import csv
import json
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from pydantic import BaseModel, Field


class FeedbackIngestionResult(BaseModel):
    source_type: str
    file_name: str
    columns: list[str] = Field(default_factory=list)
    rows: list[dict[str, Any]] = Field(default_factory=list)
    raw_text_blocks: list[str] = Field(default_factory=list)
    parse_warnings: list[str] = Field(default_factory=list)
    question_mapping: dict[str, str] = Field(default_factory=dict)
    metadata: dict[str, Any] = Field(default_factory=dict)


class FeedbackIngestionService:
    def ingest(self, file_name: str, raw_content: bytes, survey: Any | None = None) -> FeedbackIngestionResult:
        suffix = Path(file_name or "feedback").suffix.lower().lstrip(".") or "txt"
        if suffix == "csv":
            result = self._parse_csv(file_name, raw_content)
        elif suffix == "json":
            result = self._parse_json(file_name, raw_content)
        elif suffix == "xlsx":
            result = self._parse_xlsx(file_name, raw_content)
        elif suffix in {"txt", "md"}:
            result = self._parse_text(file_name, raw_content, suffix)
        else:
            result = self._parse_text(file_name, raw_content, suffix)
            result.parse_warnings.append(f"暂不支持 .{suffix} 的结构化解析，已按文本反馈处理。")
        if survey is not None:
            result.question_mapping = self._build_question_mapping(result.columns, survey)
        return result

    def _parse_csv(self, file_name: str, raw_content: bytes) -> FeedbackIngestionResult:
        text = raw_content.decode("utf-8-sig")
        reader = csv.DictReader(StringIO(text))
        rows = [{str(k): v for k, v in row.items() if k is not None} for row in reader]
        return FeedbackIngestionResult(source_type="csv", file_name=file_name, columns=list(reader.fieldnames or []), rows=rows)

    def _parse_json(self, file_name: str, raw_content: bytes) -> FeedbackIngestionResult:
        payload = json.loads(raw_content.decode("utf-8-sig"))
        if isinstance(payload, list):
            rows = [item for item in payload if isinstance(item, dict)]
        elif isinstance(payload, dict):
            data = payload.get("rows") or payload.get("responses") or payload.get("data")
            rows = [item for item in data if isinstance(item, dict)] if isinstance(data, list) else [payload]
        else:
            rows = []
        columns = sorted({str(key) for row in rows for key in row.keys()})
        return FeedbackIngestionResult(source_type="json", file_name=file_name, columns=columns, rows=rows)

    def _parse_text(self, file_name: str, raw_content: bytes, source_type: str) -> FeedbackIngestionResult:
        text = raw_content.decode("utf-8-sig", errors="replace")
        blocks = [block.strip() for block in text.split("\n\n") if block.strip()]
        if not blocks and text.strip():
            blocks = [line.strip() for line in text.splitlines() if line.strip()]
        rows = [{"feedback_text": block} for block in blocks]
        return FeedbackIngestionResult(
            source_type=source_type,
            file_name=file_name,
            columns=["feedback_text"] if rows else [],
            rows=rows,
            raw_text_blocks=blocks,
            metadata={"unstructured": True},
        )

    def _parse_xlsx(self, file_name: str, raw_content: bytes) -> FeedbackIngestionResult:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:  # noqa: BLE001
            return self._parse_xlsx_minimal(file_name, raw_content, f"openpyxl unavailable: {exc}")
        workbook = load_workbook(BytesIO(raw_content), read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header = [str(value).strip() if value is not None else "" for value in next(rows_iter, [])]
        rows = []
        for values in rows_iter:
            row = {header[index]: values[index] for index in range(min(len(header), len(values))) if header[index]}
            if any(value not in {None, ""} for value in row.values()):
                rows.append(row)
        return FeedbackIngestionResult(source_type="xlsx", file_name=file_name, columns=header, rows=rows)

    def _parse_xlsx_minimal(self, file_name: str, raw_content: bytes, warning: str) -> FeedbackIngestionResult:
        try:
            with ZipFile(BytesIO(raw_content)) as zf:
                names = zf.namelist()
        except Exception as exc:  # noqa: BLE001
            return FeedbackIngestionResult(source_type="xlsx", file_name=file_name, parse_warnings=[warning, f"xlsx parse failed: {exc}"])
        return FeedbackIngestionResult(
            source_type="xlsx",
            file_name=file_name,
            parse_warnings=[warning, "当前环境缺少 openpyxl，已确认文件为 xlsx 但未抽取表格行。"],
            metadata={"zip_entries": names[:10]},
        )

    def _build_question_mapping(self, columns: list[str], survey: Any) -> dict[str, str]:
        questions = list(getattr(survey, "questions", []) or [])
        mapping: dict[str, str] = {}
        for column in columns:
            normalized = _normalize(column)
            for question in questions:
                candidates = [
                    getattr(question, "question_id", ""),
                    getattr(question, "field_name", ""),
                    getattr(question, "question_text", ""),
                ]
                if any(normalized == _normalize(candidate) for candidate in candidates):
                    mapping[column] = getattr(question, "field_name", column)
                    break
                question_text = str(getattr(question, "question_text", ""))
                if normalized and (normalized in _normalize(question_text) or _normalize(question_text) in normalized):
                    mapping[column] = getattr(question, "field_name", column)
                    break
            mapping.setdefault(column, column)
        return mapping


def _normalize(value: Any) -> str:
    return "".join(str(value or "").lower().split())
